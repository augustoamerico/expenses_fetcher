import os
from datetime import datetime
from random import randint
from time import sleep
from typing import List, Dict, Callable

from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from selenium.common.exceptions import ElementNotInteractableException
from src.application.password_getter.password_getter import IPasswordGetter

from src.infrastructure.bank_account_transactions_fetchers.exceptions import (
    AuthenticationException,
    DownloadTransactionsXlsException,
)
from src.infrastructure.bank_account_transactions_fetchers.i_transactions_fetcher import (
    ITransactionsFetcher,
)

LOGINPAGE = "https://ind.activobank.pt/_loginV2/BlueMainLoginCdm.aspx?ReturnUrl=https%3a%2f%2find.activobank.pt%2fpt%2fprivate%2fdia-a-dia%2fPages%2fdia-a-dia.aspx"
INPUTCODEIDS = [
    "BlueMainLoginControlCdm1_txt_1_position",
    "BlueMainLoginControlCdm1_txt_2_position",
    "BlueMainLoginControlCdm1_txt_3_position",
]
XLSX_PROPS = {
    "MAIN": {
        "start_row": 8,
        "start_column": "A",
        "end_column": "E",
        "expected_header": ["Data Lanc.", "Data Valor", "Descrição", "Valor", "Saldo"],
    },
    "PRE": {
        "start_row": 7,
        "start_column": "A",
        "end_column": "D",
        "expected_header": ["Data Lanc.", "Data Valor", "Descrição", "Valor"],
    },
}


class ActiveBankCardCrawler(ITransactionsFetcher):
    def __init__(
        self,
        card_number: str,
        driver,
        driver_waiter,
        tmp_download_dir: str,
        xlsx_props,
        sidebar_main_entry_id: str,
        sidebar_menu_entry_cards_id: str,
        search_start_date_element_id: str,
        search_end_date_element_id: str,
        search_button_element_id: str,
        download_excel_button_id: str,
    ):
        self.card_number = card_number
        self.driver = driver
        self.driver_waiter = driver_waiter
        self.tmp_download_dir = tmp_download_dir
        self.xlsx_props = xlsx_props
        self.sidebar_main_entry_id = sidebar_main_entry_id
        self.sidebar_menu_entry_cards_id = sidebar_menu_entry_cards_id
        self.search_start_date_element_id = search_start_date_element_id
        self.search_end_date_element_id = search_end_date_element_id
        self.search_button_element_id = search_button_element_id
        self.download_excel_button_id = download_excel_button_id

    def getTransactions(
        self, date_init: datetime = None, date_end: datetime = None
    ) -> List[Dict[str, object]]:
        try:
            # assume the page is the one after login successfully
            self._download_transactions_xlsx(date_init, date_end)

            tmp_folder_content = os.listdir(self.tmp_download_dir)
            transactions = []

            if len(tmp_folder_content) == 1:
                file_path = tmp_folder_content.pop()
                transactions = self._get_trx_from_xlsx(file_path, self.xlsx_props)
            else:
                raise DownloadTransactionsXlsException("Something went wrong.")

            return transactions
        except Exception as e:
            self.driver.quit()
            raise e

    def _get_trx_from_xlsx(
        self, file_path, xls_props: Dict[str, object]
    ) -> List[Dict[str, object]]:
        file_path = os.path.join(self.tmp_download_dir, file_path)
        tmp_trx_xlsx = load_workbook(file_path)
        tmp_trx_sheet = tmp_trx_xlsx.worksheets[0]
        max_row = tmp_trx_sheet.max_row

        header = [
            header_col.value
            for header_col in tmp_trx_sheet[
                f"{xls_props['start_column']}"
                f"{xls_props['start_row']}:"
                f"{xls_props['end_column']}"
                f"{xls_props['start_row']}"
            ][0]
        ]
        assert header == xls_props["expected_header"]
        rows = []

        for row in tmp_trx_sheet[
            f"{xls_props['start_column']}{xls_props['start_row'] + 1}:"
            f"{xls_props['end_column']}{max_row}"
        ]:
            rows.append(dict(zip(header, [column.value for column in row])))

        os.remove(file_path)
        return rows

    def _download_transactions_xlsx(self, date_init: datetime, date_end: datetime):
        self.driver.find_element_by_id(self.sidebar_main_entry_id).click()
        self.driver_waiter.until(
            EC.presence_of_element_located((By.ID, self.search_button_element_id))
        )

        card_element = self.driver.find_element_by_id(
            self.sidebar_menu_entry_cards_id
        ).find_elements_by_xpath(f"//*[contains(text(), '{self.card_number}')]")[0]
        card_element.click()

        self.driver_waiter.until(
            EC.presence_of_element_located((By.ID, self.search_start_date_element_id))
        )
        sleep(2)

        start_date = self.driver.find_element_by_id(self.search_start_date_element_id)
        start_date.clear()
        start_date.send_keys(date_init.strftime("%d/%m/%Y"))
        # date_init_str = date_init.strftime('%d/%m/%Y')
        # self.driver.execute_script(f"document.getElementById('{self.search_start_date_element_id}').value='{date_init_str}'")
        sleep(2)
        end_date = self.driver.find_element_by_id(self.search_end_date_element_id)
        end_date.clear()
        end_date.send_keys(date_end.strftime("%d/%m/%Y"))
        # sleep(3)
        # self.driver.find_element_by_id(self.search_button_element_id).click()
        clicked_on_search = self._try_action_for_next_seconds(
            lambda: self.driver.find_element_by_id(
                self.search_button_element_id
            ).click(),
            5,
        )
        if not clicked_on_search:
            raise Exception("Couldn't search transactions")
        self.driver_waiter.until(
            EC.presence_of_element_located((By.ID, "divShowWaitPanel"))
        )
        self.driver_waiter.until_not(
            EC.presence_of_element_located((By.ID, "divShowWaitPanel"))
        )
        sleep(10)
        clicked_on_download = self._try_action_for_next_seconds(
            lambda: self.driver.find_element_by_id(
                self.download_excel_button_id
            ).click(),
            10,
        )
        if not clicked_on_download:
            raise Exception("Couldn't fetch transactions")
        sleep(randint(1, 3))

    def _try_action_for_next_seconds(
        self, action: Callable, seconds, wait_for_seconds=1
    ):

        for i in range(int(seconds / wait_for_seconds)):
            try:
                action()
                return True
            except:
                sleep(wait_for_seconds)
        return False


class ActiveBankMainCard(ActiveBankCardCrawler):
    def __init__(self, card_number, driver, driver_waiter, tmp_download_dir):
        super().__init__(
            card_number,
            driver,
            driver_waiter,
            tmp_download_dir,
            XLSX_PROPS["MAIN"],
            "ctl00_PlaceHolderMain_DayInfo1__leftMenu1_lbtAccounts",
            "itemListAccount",
            "ctl02_searchStartDate",
            "ctl02_endStartDate",
            "searchMovs",
            "ctl02__lnkExcelExport",
        )


class ActiveBankPreCard(ActiveBankCardCrawler):
    def __init__(self, card_number, driver, driver_waiter, tmp_download_dir):
        super().__init__(
            card_number,
            driver,
            driver_waiter,
            tmp_download_dir,
            XLSX_PROPS["PRE"],
            "ctl00_PlaceHolderMain_DayInfo1__leftMenu1_lbtCards",
            "itemListCards",
            "ctl02_searchStartDate",
            "ctl02_searchEndDate",
            "lbtSearch",
            "ctl02__lnkExcelExport",
        )


class ActiveBankCrawler:
    def __init__(
        self,
        username: str,
        password: str,
        tmp_download_dir: str,
        get_password: Callable,
    ):
        self.tmp_download_dir = os.path.join(
            tmp_download_dir, str(datetime.now().strftime("%Y_%m_%d_%H_%M_%S"))
        )
        self.username = username
        self.password = password
        self.trx = None
        self.cards: Dict[str, ActiveBankCardCrawler] = dict()
        self.get_password = get_password

        os.makedirs(self.tmp_download_dir, exist_ok=True)

        prefs = dict()
        prefs["profile.default_content_settings.popups"] = 0
        prefs["download.default_directory"] = self.tmp_download_dir
        prefs["browser.tabs.warnOnClose"] = False
        opts = webdriver.ChromeOptions()
        opts.add_experimental_option("prefs", prefs)
        opts.add_argument("--headless")
        self.driver = webdriver.Chrome(chrome_options=opts)
        self.driver_waiter = WebDriverWait(self.driver, timeout=30)
        try:
            self.driver.get(LOGINPAGE)
            self.driver_waiter.until(
                EC.presence_of_element_located((By.ID, "divBtnShort"))
            )
            self._log_in()
            self.cards = self.__get_cards()
        except Exception as e:
            self.driver.quit()
            raise e

    def __get_cards(self) -> Dict[str, Callable]:
        # element get main card ctl00_PlaceHolderMain_DayInfo1__leftMenu1__ddaAccList_rptAccounts_ctl01_lnkTable
        # Main account
        cards_to_fetcher = {}
        self._try_action_for_next_seconds(
            lambda: self.driver.find_element_by_id(
                "ctl00_PlaceHolderMain_DayInfo1__leftMenu1_lbtAccounts"
            ).click(),
            5,
            0.5,
        )
        self.driver_waiter.until(EC.presence_of_element_located((By.ID, "searchMovs")))
        for card_element in self.driver.find_element_by_id(
            "itemListAccount"
        ).find_elements_by_css_selector("ul>div>li>a>span[class='title']"):
            card_number = card_element.text.strip()
            cards_to_fetcher[card_number] = ActiveBankMainCard(
                card_number, self.driver, self.driver_waiter, self.tmp_download_dir
            )

        self._try_action_for_next_seconds(
            lambda: self.driver.find_element_by_id(
                "ctl00_PlaceHolderMain_DayInfo1__leftMenu1_lbtCards"
            ).click(),
            5,
            0.5,
        )
        self.driver_waiter.until(EC.presence_of_element_located((By.ID, "lbtSearch")))
        for card_element in self.driver.find_element_by_id(
            "itemListCards"
        ).find_elements_by_css_selector("ul>div>li>a>span[class='title']"):
            card_number = card_element.text.strip()
            cards_to_fetcher[card_number] = ActiveBankPreCard(
                card_number, self.driver, self.driver_waiter, self.tmp_download_dir
            )

        return cards_to_fetcher

    def get_card(self, card_number: str) -> ITransactionsFetcher:
        return self.cards.get(card_number, None)

    def _set_username(self):
        element = self.driver.find_element_by_name(
            "BlueMainLoginControlCdm1$txtUserCode"
        )
        element.clear()
        element.send_keys(self.username)
        element.send_keys(Keys.RETURN)
        self.driver_waiter.until(EC.presence_of_element_located((By.ID, "divBtnLogOn")))

    def _banksite_user_set_password(self):
        codes = self.driver.find_elements_by_xpath(
            '//div[@id="BlueMainLoginControlCdm1_divCode"]//ul//li[not(@type="password")]'
        )

        codeDigits = []
        for codeBox in codes:
            try:
                codeDigits.append(int(codeBox.text[0]))
            except:
                continue
        for idx, code_position in enumerate(codeDigits):
            elemInput = self.driver.find_element_by_id(INPUTCODEIDS[idx])
            elemInput.clear()
            elemInput.send_keys(self.password[code_position - 1])
        self.driver.find_element_by_id("divBtnLogOn").click()
        self.driver_waiter.until(
            EC.presence_of_element_located((By.ID, "_lnkBtnConfirm"))
        )
        self.driver.find_element_by_id("_lnkBtnConfirm").click()
        self.driver_waiter.until(
            EC.presence_of_element_located(
                (By.ID, "ctl00_PlaceHolderMain_DayInfo1__leftMenu1_lbtAccounts")
            )
        )

    def _set_password(self):
        if self.password is None:
            self.password = self.get_password()
        code_digits = []
        codes = self.driver.find_elements_by_xpath(
            '//div[@id="BlueMainLoginControlCdm1_divCode"]//ul//li[not(@type="password")]'
        )
        for codeBox in codes:
            try:
                code_digits.append(int(codeBox.text[0]))
            except:
                continue
        for x in range(0, len(code_digits)):
            elemInput = self.driver.find_element_by_id(INPUTCODEIDS[x])
            elemInput.clear()
            elemInput.send_keys(self.password[code_digits[x] - 1])
        self.driver.find_element_by_id("divBtnLogOn").click()
        sleep(5)
        self.driver.find_element_by_id("_lnkBtnConfirm").click()

    def _log_in(self):

        if self.username is None:
            raise AuthenticationException(
                "Need at least the user to access bank account"
            )

        self._set_username()
        self._set_password()

        try:
            self._try_action_for_next_seconds(
                lambda: self.driver.find_elements_by_css_selector(
                    ".cookieButtonV2 a[onclick='SetCookies();']"
                )[0].click(),
                5,
            )
        except ElementNotInteractableException:
            pass

    def _try_action_for_next_seconds(
        self, action: Callable, seconds, wait_for_seconds=1
    ):
        last_exception = None

        for i in range(int(seconds / wait_for_seconds)):
            try:
                action()
                return True
            except Exception as e:
                last_exception = e
                sleep(wait_for_seconds)
        if last_exception is not None:
            raise last_exception
        return False
