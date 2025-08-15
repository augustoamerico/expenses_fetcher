from typing import List, Dict, Optional
from datetime import datetime
from openpyxl import load_workbook

from src.infrastructure.bank_account_transactions_fetchers.i_transactions_fetcher import (
    ITransactionsFetcher,
)


class XlsxTransactionsFetcher(ITransactionsFetcher):
    def __init__(
        self,
        header_skip_rows: int,
        date_format: str,
        decimal_separator: str,
        thousands_separator: str,
        columns: Dict[str, str],
        sheet_name: Optional[str] = None,
        footer_skip_rows: int = 0,
    ):
        self.header_skip_rows = header_skip_rows
        self.date_format = date_format
        self.decimal_separator = decimal_separator
        self.thousands_separator = thousands_separator
        self.columns = columns
        self.sheet_name = sheet_name
        self.footer_skip_rows = footer_skip_rows

    def _parse_date(self, value) -> Optional[datetime]:
        try:
            if value is None or value == "":
                return None
            if isinstance(value, datetime):
                return value
            # treat as string
            return datetime.strptime(str(value).strip(), self.date_format)
        except Exception:
            # Non-date footer or malformed cell
            return None

    def _parse_amount(self, value: Optional[str]) -> float:
        if value is None or value == "":
            return 0.0
        s = str(value).strip()
        if self.thousands_separator:
            s = s.replace(self.thousands_separator, "")
        if self.decimal_separator and self.decimal_separator != ".":
            s = s.replace(self.decimal_separator, ".")
        return float(s)

    def getTransactions(
        self, date_init: datetime = None, date_end: datetime = None, file_path: str = None
    ) -> List[Dict[str, object]]:
        if not file_path:
            raise Exception("file_path must be provided to XlsxTransactionsFetcher")
        wb = load_workbook(filename=file_path, data_only=True)
        ws = wb[self.sheet_name] if self.sheet_name else wb.worksheets[0]

        # Build header index from the first non-skipped row
        header_row_idx = self.header_skip_rows + 1
        headers = [c.value for c in ws[header_row_idx]]
        col_idx = {h: i for i, h in enumerate(headers)}

        def idx(name: str) -> int:
            if name not in col_idx:
                raise Exception(f"Column '{name}' not found in XLSX headers")
            return col_idx[name]

        # Determine last data row (exclude footer rows if configured)
        max_row_inclusive = ws.max_row
        if self.footer_skip_rows and max_row_inclusive - self.footer_skip_rows > header_row_idx:
            max_row_inclusive = max_row_inclusive - self.footer_skip_rows

        rows = []
        for row in ws.iter_rows(min_row=header_row_idx + 1, max_row=max_row_inclusive, values_only=True):
            capture = self._parse_date(row[idx(self.columns["capture_date"])])
            auth = self._parse_date(row[idx(self.columns.get("auth_date", self.columns["capture_date"]))])
            if auth is None:
                auth = capture

            # Skip rows without any valid date (likely footers or banners)
            if capture is None and auth is None:
                continue

            description_cell = row[idx(self.columns["description"])]
            description = str(description_cell).strip() if description_cell is not None else ""
            debit = self._parse_amount(row[idx(self.columns["debit"])])
            credit = self._parse_amount(row[idx(self.columns["credit"])])
            amount = credit - debit
            balance = None
            if "balance" in self.columns:
                try:
                    balance = self._parse_amount(row[idx(self.columns["balance"])])
                except Exception:
                    balance = None

            # Date range filter
            if date_init is not None and auth.date() < date_init.date():
                continue
            if date_end is not None and auth.date() > date_end.date():
                continue

            rows.append(
                {
                    "captureDate": capture,
                    "authDate": auth,
                    "description": description,
                    "amount": amount,
                    "balance": balance,
                }
            )

        return rows
